import pandas as pd
import sqlite3
import os
import sys
import glob
from datetime import datetime, timedelta

# 設定: ファイルパス
BASE_DIR = r"c:\分析"
DB_FILE = os.path.join(BASE_DIR, "fjallraven_md.db")
OUTPUT_EXCEL = os.path.join(BASE_DIR, f"MD分析レポート_{datetime.now().strftime('%Y%m%d')}.xlsx")
LOG_FILE = os.path.join(BASE_DIR, "process_log.txt")

def log(msg):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] {msg}\n")
    print(msg)

def find_latest_file(pattern):
    files = glob.glob(os.path.join(BASE_DIR, pattern))
    if not files:
        raise Exception(f"ファイルが見つかりません: {pattern}")
    # 更新日時が一番新しいファイルを取得
    latest_file = max(files, key=os.path.getmtime)
    log(f"自動選択: {os.path.basename(latest_file)}")
    return latest_file

def load_csv(path):
    log(f"読み込み中: {os.path.basename(path)}")
    for enc in ['cp932', 'utf-8-sig', 'utf-8']:
        try:
            return pd.read_csv(path, encoding=enc)
        except Exception:
            continue
    raise Exception(f"ファイルの読み込みに失敗しました: {path}")

def main():
    try:
        if os.path.exists(LOG_FILE):
            os.remove(LOG_FILE)
        
        log("処理を開始します。")
        
        try:
            import openpyxl
        except ImportError:
            log("エラー: Excel出力に必要なライブラリ 'openpyxl' が見つかりません。")
            return
        
        # 1. ファイルの柔軟な読み込み（ゆらぎ対応）
        sales_file = find_latest_file("*売上*.csv")
        stock_file = find_latest_file("*在庫*.csv")
        master_file = find_latest_file("*マスタ*.csv")
        
        df_sales = load_csv(sales_file)
        df_stock = load_csv(stock_file)
        df_master = load_csv(master_file)
        
        log(f"売上: {len(df_sales)}件, 在庫: {len(df_stock)}件, マスタ: {len(df_master)}件")

        # カラム名のクリーンアップ
        df_sales.columns = [c.strip() for c in df_sales.columns]
        df_stock.columns = [c.strip() for c in df_stock.columns]
        df_master.columns = [c.strip() for c in df_master.columns]
        
        # マスタカラム: Name(C=2), Brand(G=6), L(M=12), M(N=13), S(O=14), Color(U=20)
        master_cols = df_master.columns.tolist()
        col_m_name = master_cols[2] if len(master_cols) > 2 else "商品名"
        col_m_brand = master_cols[6] if len(master_cols) > 6 else "ブランド"
        col_m_cat_l = master_cols[12] if len(master_cols) > 12 else "大分類"
        col_m_cat_m = master_cols[13] if len(master_cols) > 13 else "中分類"
        col_m_cat_s = master_cols[14] if len(master_cols) > 14 else "小分類"
        col_m_color = master_cols[20] if len(master_cols) > 20 else "カラー"
        
        # 売上カラム: Reg(A=0), Size(J=9), Price(O=14), Inbound(U=20), Age(V=21), Member(Y=24)
        sales_cols = df_sales.columns.tolist()
        col_s_reg = sales_cols[0]
        col_s_size = sales_cols[9] if len(sales_cols) > 9 else "サイズ"
        col_s_price = sales_cols[14] if len(sales_cols) > 14 else None
        col_s_inbound = sales_cols[20] if len(sales_cols) > 20 else None
        col_s_age = sales_cols[21] if len(sales_cols) > 21 else None
        col_s_member = sales_cols[24] if len(sales_cols) > 24 else None
        
        # 結合キー設定
        try:
            col_s_key = df_sales.columns[3] # D列
            col_i_key = df_stock.columns[2] # C列
            col_m_key = df_master.columns[0] # A列
        except IndexError as e:
            log(f"エラー: 指定された列が見つかりません。({str(e)})")
            return
        
        df_sales[col_s_key] = df_sales[col_s_key].astype(str)
        df_stock[col_i_key] = df_stock[col_i_key].astype(str)
        df_master[col_m_key] = df_master[col_m_key].astype(str)
        
        # 売上＋マスタ結合
        df_merged = pd.merge(df_sales, df_master, left_on=col_s_key, right_on=col_m_key, how='left', suffixes=('', '_master'))
        
        for attr in ['col_m_name', 'col_m_brand', 'col_m_cat_l', 'col_m_cat_m', 'col_m_cat_s', 'col_m_color']:
            val = locals()[attr]
            if val not in df_merged.columns and f"{val}_master" in df_merged.columns:
                locals()[attr] = f"{val}_master"
        
        # 在庫集計 (品番レベル)
        stock_val_col = df_stock.columns[-1]
        for c in df_stock.columns:
            if any(x in c for x in ['在庫', '数量', '残数']):
                stock_val_col = c
                break
        df_stock_sum = df_stock.groupby(col_i_key)[stock_val_col].sum().reset_index()
        
        # 統合データに在庫を付与
        df_merged = pd.merge(df_merged, df_stock_sum, left_on=col_s_key, right_on=col_i_key, how='left')
        df_merged[stock_val_col] = df_merged[stock_val_col].fillna(0)
        
        # 日付と数値のクリーンアップ
        col_date = None
        for c in df_sales.columns:
            if any(x in c for x in ['日', '日付', '年月日', 'Date']):
                col_date = c
                break
        if col_date:
            df_merged[col_date] = pd.to_datetime(df_merged[col_date], errors='coerce')
        
        if col_s_price:
            df_merged[col_s_price] = df_merged[col_s_price].astype(str).str.replace('¥', '').str.replace(',', '').str.strip()
            df_merged[col_s_price] = pd.to_numeric(df_merged[col_s_price], errors='coerce').fillna(0)
            
        col_m_cost = master_cols[11] if len(master_cols) > 11 else None
        if col_m_cost:
            df_merged[col_m_cost] = df_merged[col_m_cost].astype(str).str.replace('¥', '').str.replace(',', '').str.strip()
            df_merged[col_m_cost] = pd.to_numeric(df_merged[col_m_cost], errors='coerce').fillna(0)

        # WOS（在庫週数）の計算 (直近14日間の平均売上に基づく)
        if col_date:
            max_date = df_merged[col_date].max()
            if pd.notnull(max_date):
                two_weeks_ago = max_date - timedelta(days=14)
                df_2w_sales = df_merged[df_merged[col_date] >= two_weeks_ago]
                sales_by_item = df_2w_sales.groupby(col_s_key)[col_s_reg].count().reset_index(name='sales_2w')
                sales_by_item['weekly_avg'] = sales_by_item['sales_2w'] / 2.0
                
                df_merged = pd.merge(df_merged, sales_by_item, on=col_s_key, how='left')
                df_merged['weekly_avg'] = df_merged['weekly_avg'].fillna(0.01)
                df_merged['WOS(在庫週数)'] = df_merged[stock_val_col] / df_merged['weekly_avg']
                df_merged['WOS(在庫週数)'] = df_merged['WOS(在庫週数)'].round(1)
            else:
                df_merged['WOS(在庫週数)'] = "N/A"
        else:
            df_merged['WOS(在庫週数)'] = "N/A"

        # --- ABC/XYZ分析の計算 ---
        # ABC分析 (売上実績ベース)
        abc_df = df_merged.groupby([col_s_key, col_m_name, col_m_cat_l]).agg(
            販売数=(col_s_reg, 'count'),
            売上金額=(col_s_price, 'sum'),
            原価合計=(col_m_cost, 'sum') if col_m_cost else (col_s_reg, lambda x: 0),
            平均在庫数=(stock_val_col, 'mean')
        ).reset_index()
        
        # 売価（平均）を逆算
        abc_df['売価'] = (abc_df['売上金額'] / abc_df['販売数']).fillna(0).round(0)
        
        # 粗利・原価率
        if col_m_cost:
            abc_df['粗利額'] = abc_df['売上金額'] - abc_df['原価合計']
            abc_df['原価率'] = (abc_df['原価合計'] / abc_df['売上金額']).fillna(0)
        else:
            abc_df['粗利額'] = 0
            abc_df['原価率'] = 0
            
        # 回転率
        abc_df['回転率'] = (abc_df['販売数'] / abc_df['平均在庫数'].replace(0, 0.01)).round(2)
        
        # 売上構成比の計算（売上順にソート）
        abc_df = abc_df.sort_values(by='売上金額', ascending=False)
        total_sales_amt = abc_df['売上金額'].sum()
        abc_df['売上構成'] = (abc_df['売上金額'] / total_sales_amt).fillna(0)
        abc_df['累計売上構成'] = abc_df['売上構成'].cumsum()
        
        # ABCランク判定 (A: ~80%, B: ~90%, C: ~100%)
        def assign_abc(cum_ratio):
            if cum_ratio <= 0.80: return 'A'
            elif cum_ratio <= 0.90: return 'B'
            else: return 'C'
        abc_df['ABCランク'] = abc_df['累計売上構成'].apply(assign_abc)

        # XYZ分析 (需要変動係数ベース)
        if col_date and pd.notnull(max_date):
            # 日別の販売数をピボット（売れなかった日も0で埋める）
            daily_pivot = df_merged.pivot_table(index=col_s_key, columns=pd.Grouper(key=col_date, freq='D'), values=col_s_reg, aggfunc='count', fill_value=0)
            
            xyz_stats = pd.DataFrame()
            xyz_stats['mean'] = daily_pivot.mean(axis=1)
            xyz_stats['std'] = daily_pivot.std(axis=1)
            xyz_stats['需要変動係数'] = (xyz_stats['std'] / xyz_stats['mean'].replace(0, 0.01)).fillna(0).round(2)
            
            def assign_xyz(cv):
                if cv <= 0.3: return 'X'
                elif cv <= 0.8: return 'Y'
                else: return 'Z'
            xyz_stats['XYZランク'] = xyz_stats['需要変動係数'].apply(assign_xyz)
            
            # ABCとXYZを結合
            abc_df = pd.merge(abc_df, xyz_stats[['需要変動係数', 'XYZランク']], on=col_s_key, how='left')
        else:
            abc_df['需要変動係数'] = 0
            abc_df['XYZランク'] = '-'
            
        abc_df['ABC×XYZ'] = abc_df['ABCランク'] + abc_df['XYZランク']

        # ABC/XYZシート用のカラム整理
        abc_xyz_output = abc_df[[
            col_s_key, col_m_name, col_m_cat_l, '売価', '販売数', '売上金額', 
            '原価率', '粗利額', '平均在庫数', '回転率', 
            '売上構成', '累計売上構成', 'ABCランク', 
            '需要変動係数', 'XYZランク', 'ABC×XYZ'
        ]].rename(columns={col_s_key: '商品コード', col_m_name: '商品名', col_m_cat_l: 'カテゴリ'})

        # --- KPI計算 ---
        total_sales_count = len(df_sales)
        total_revenue = df_merged[col_s_price].sum() if col_s_price else 0
        unique_transactions = df_merged[col_s_reg].nunique()
        avg_spend = total_revenue / unique_transactions if unique_transactions > 0 else 0
        items_per_trans = total_sales_count / unique_transactions if unique_transactions > 0 else 0
        
        member_ratio = df_merged[col_s_member].notnull().mean() if col_s_member else 0
        inbound_ratio = df_merged[col_s_inbound].astype(str).str.contains('外国', na=False).mean() if col_s_inbound else 0
        age_dist = df_merged[col_s_age].value_counts(normalize=True).to_frame(name='構成比') if col_s_age else pd.DataFrame()

        cat_summary = df_merged.groupby(col_m_cat_l).agg({col_s_reg: 'count'}).rename(columns={col_s_reg: '売上数'})
        cat_summary['構成比'] = cat_summary['売上数'] / cat_summary['売上数'].sum()

        top10 = df_merged.groupby([col_m_cat_l, col_m_name]).agg({col_s_reg: 'count'}).sort_values([col_m_cat_l, col_s_reg], ascending=[True, False])
        top10 = top10.groupby(level=0).head(10)
        
        trend_2w = None
        if col_date and pd.notnull(max_date):
            trend_2w = df_2w_sales.groupby([pd.Grouper(key=col_date, freq='D'), col_m_cat_l]).size().unstack().fillna(0)

        # 2. データベース蓄積対応 (Append)
        log("データベースを更新中...")
        conn = sqlite3.connect(DB_FILE)
        
        # 在庫データにはスナップショット日を付与して保存
        df_stock_save = df_stock.copy()
        df_stock_save['実行日'] = datetime.now().strftime('%Y-%m-%d')
        df_stock_save.to_sql('stock_history', conn, if_exists='append', index=False)
        
        # 売上データ（統合版）は、日付で重複を防ぎつつAppend
        if col_date and pd.notnull(df_merged[col_date].min()):
            min_date = df_merged[col_date].min().strftime('%Y-%m-%d %H:%M:%S')
            max_date = df_merged[col_date].max().strftime('%Y-%m-%d %H:%M:%S')
            try:
                # 同じ期間のデータが既にあれば削除（重複防止）
                cursor = conn.cursor()
                cursor.execute(f'DELETE FROM sales_analysis WHERE "{col_date}" BETWEEN ? AND ?', (min_date, max_date))
                conn.commit()
            except sqlite3.OperationalError:
                pass # テーブルが存在しない場合は無視
        
        # 文字列型にしてから保存（DBエラー回避のため）
        df_db_save = df_merged.copy()
        if col_date:
            df_db_save[col_date] = df_db_save[col_date].astype(str)
            
        # WOS計算用に追加した列はDBのスキーマに合わないため除外して保存
        df_db_save = df_db_save.drop(columns=['sales_2w', 'weekly_avg', 'WOS(在庫週数)'], errors='ignore')
        df_db_save.to_sql('sales_analysis', conn, if_exists='append', index=False)
        df_master.to_sql('master_data', conn, if_exists='replace', index=False) # マスタは最新を維持
        conn.close()

        # 3. Excel出力
        summary_df = pd.DataFrame({
            '項目': ['全体の売上個数', '客数 (買上点数)', '合計売上高 (税抜)', '客単価', 'セット率', 'インバウンド比率', '会員比率', '最終データ更新日'],
            '値': [total_sales_count, unique_transactions, f"¥{total_revenue:,.0f}", f"¥{avg_spend:,.0f}", f"{items_per_trans:.2f}", f"{inbound_ratio:.0%}", f"{member_ratio:.0%}", datetime.now().strftime('%Y-%m-%d')]
        })

        # カテゴリー別の在庫集計（中分類ベースに変更）
        cat_summary = df_merged.groupby(col_m_cat_m).agg({col_s_reg: 'count'}).rename(columns={col_s_reg: '売上数'})
        cat_summary['構成比'] = cat_summary['売上数'] / cat_summary['売上数'].sum()
        
        cat_stock = df_merged.drop_duplicates(subset=[col_s_key]).groupby(col_m_cat_m)[stock_val_col].sum()
        cat_summary = pd.merge(cat_summary, cat_stock, on=col_m_cat_m, how='left').rename(columns={stock_val_col: '在庫数'})
        cat_summary['在庫構成'] = cat_summary['在庫数'] / cat_summary['在庫数'].sum()

        log("Excelレポートを生成中（可視化処理含む）...")
        try:
            from openpyxl.styles import PatternFill, Color
            from openpyxl.formatting.rule import CellIsRule, DataBarRule
            from openpyxl.chart import BarChart, Reference
            from openpyxl.chart.label import DataLabelList
            
            with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
                summary_df.to_excel(writer, sheet_name='総合サマリー', index=False)
                cat_summary.to_excel(writer, sheet_name='カテゴリー構成比')
                top10.to_excel(writer, sheet_name='全店TOP10')
                # 直近1ヶ月のトレンド分析 (1M分析)
                if col_date and pd.notnull(df_merged[col_date].max()):
                    current_max_date = df_merged[col_date].max()
                    one_month_ago = current_max_date - timedelta(days=30)
                    df_1m_sales = df_merged[df_merged[col_date] >= one_month_ago]
                    trend_1m = df_1m_sales.groupby([pd.Grouper(key=col_date, freq='D'), col_m_cat_m]).size().unstack().fillna(0)
                    trend_1m.to_excel(writer, sheet_name='直近1ヶ月トレンド')
                    
                    # 1M分析シートの装飾（ヒートマップ）
                    ws_1m = writer.sheets['直近1ヶ月トレンド']
                    from openpyxl.formatting.rule import ColorScaleRule
                    # 白から赤へのカラースケール（売上が多いほど赤く）
                    color_scale_rule = ColorScaleRule(start_type='min', start_color='FFFFFF',
                                                      end_type='max', end_color='FF6347')
                    # 30日分＋アルファをカバー（B2:AZ40）
                    ws_1m.conditional_formatting.add('B2:AZ40', color_scale_rule)

                abc_xyz_output.to_excel(writer, sheet_name='ABC_XYZ分析', index=False, startrow=11)
                
                # ABC_XYZ分析シートの装飾
                ws_abc = writer.sheets['ABC_XYZ分析']
                ws_abc['A1'] = '【ABC×XYZランクの見方】'
                ws_abc['A2'] = 'AX: 超主力・安定（欠品厳禁・最優先確保）'
                ws_abc['A3'] = 'AY: 主力・中変動（在庫維持・安定供給）'
                ws_abc['A4'] = 'AZ: 主力・不安定（波あり・機動的発注）'
                ws_abc['A5'] = 'BX: 準主力・安定（定番品・在庫効率重視）'
                ws_abc['A6'] = 'BY: 準主力・中変動（動向注視・適正在庫）'
                ws_abc['A7'] = 'BZ: 準主力・不安定（販促または処分検討）'
                ws_abc['A8'] = 'CX: 下位・安定（ロングテール・低在庫維持）'
                ws_abc['A9'] = 'CY: 下位・中変動（需要がある限り維持）'
                ws_abc['A10'] = 'CZ: 下位・不安定（非効率・取扱終了の検討）'
                ws_abc.freeze_panes = 'A13'
                
                # 条件付き書式
                data_bar_rule = DataBarRule(start_type='min', end_type='max', color="638EC6", showValue=True, minLength=None, maxLength=None)
                ws_abc.conditional_formatting.add('J13:J1000', data_bar_rule)
                green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                ws_abc.conditional_formatting.add('P13:P1000', CellIsRule(operator='equal', formula=['"AX"'], fill=green_fill))
                ws_abc.conditional_formatting.add('P13:P1000', CellIsRule(operator='equal', formula=['"CZ"'], fill=red_fill))

                # ABC_XYZ分析シートの数値を％表示に設定 (G列:原価率, K列:売上構成, L列:累計売上構成)
                for col_idx in [7, 11, 12]:
                    for row in ws_abc.iter_rows(min_row=13, min_col=col_idx, max_col=col_idx):
                        for cell in row: cell.number_format = '0%'

                if not age_dist.empty:
                    age_dist.to_excel(writer, sheet_name='年代別構成')
                
                # 詳細データ
                col_s_store = sales_cols[2] if len(sales_cols) > 2 else None
                detail_cols = [col_date, col_s_store, col_s_key, col_m_name, col_m_color, col_s_size, col_m_cat_l, col_s_price, stock_val_col, 'WOS(在庫週数)']
                detail_cols = [c for c in detail_cols if c and c in df_merged.columns]
                df_merged[detail_cols].head(5000).to_excel(writer, sheet_name='売上在庫詳細(TOP5000)', index=False)
                
                ws_detail = writer.sheets['売上在庫詳細(TOP5000)']
                ws_detail.conditional_formatting.add('J2:J5001', CellIsRule(operator='lessThan', formula=['2'], fill=red_fill))
                yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                ws_detail.conditional_formatting.add('J2:J5001', CellIsRule(operator='greaterThan', formula=['20'], fill=yellow_fill))

                # カテゴリー構成比シートにグラフ追加（中分類ベース）
                ws_cat = writer.sheets['カテゴリー構成比']
                
                # 数値を小数点なしの％表示に設定
                for row in ws_cat.iter_rows(min_row=2, min_col=3, max_col=3): # C列: 構成比
                    for cell in row: cell.number_format = '0%'
                for row in ws_cat.iter_rows(min_row=2, min_col=5, max_col=5): # E列: 在庫構成
                    for cell in row: cell.number_format = '0%'

                chart = BarChart()
                chart.type = "bar"
                chart.style = 11
                chart.title = "中分類別 売上vs在庫 バランス（構成比）"
                chart.x_axis.title = '中分類'
                chart.y_axis.title = '構成比'
                
                # データラベルの追加と％形式設定
                chart.dLbls = DataLabelList()
                chart.dLbls.showVal = True
                chart.dLbls.numFmt = '0%' # グラフ上も小数点なし％表示に
                
                sales_ref = Reference(ws_cat, min_col=3, min_row=1, max_row=len(cat_summary)+1)
                chart.add_data(sales_ref, titles_from_data=True)
                stock_ref = Reference(ws_cat, min_col=5, min_row=1, max_row=len(cat_summary)+1)
                chart.add_data(stock_ref, titles_from_data=True)
                
                cats = Reference(ws_cat, min_col=1, min_row=2, max_row=len(cat_summary)+1)
                chart.set_categories(cats)
                chart.height = 20 # 中分類は数が多いのでさらに縦長に
                chart.width = 20
                ws_cat.add_chart(chart, "G2")

            
        except PermissionError:
            log("--------------------------------------------------")
            log("【重要】エラー: Excelファイルを書き込めません！")
            log("Excelを閉じてから、もう一度実行してください。")
            log("--------------------------------------------------")
            return
            
        log(f"完了しました。レポート: {os.path.basename(OUTPUT_EXCEL)}")
        
    except Exception as e:
        log(f"致命的なエラーが発生しました: {str(e)}")
        import traceback
        log(traceback.format_exc())

if __name__ == "__main__":
    main()
