import pandas as pd
import sqlite3
import os
import sys
import glob
from datetime import datetime, timedelta

# =====================================
# 設定
# =====================================
BASE_DIR   = r"c:\分析"
V2_DIR     = r"c:\分析\v2"
DB_FILE    = os.path.join(V2_DIR, "fjallraven_md_v2.db")
OUTPUT_EXCEL = os.path.join(V2_DIR, f"MD分析レポートv2_{datetime.now().strftime('%Y%m%d')}.xlsx")
LOG_FILE   = os.path.join(V2_DIR, "process_log.txt")

# 店舗リスト（閉店含む）
CLOSED_STORES = ["心斎橋パルコ"]

# =====================================
# ユーティリティ
# =====================================
def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"[{ts}] {msg}\n")
    print(msg)

def find_latest_file(pattern):
    files = glob.glob(os.path.join(BASE_DIR, pattern))
    if not files:
        raise FileNotFoundError(f"ファイルが見つかりません: {pattern}")
    latest = max(files, key=os.path.getmtime)
    log(f"自動選択: {os.path.basename(latest)}")
    return latest

def find_file_by_keyword(keyword):
    """キーワードを含む特定ファイルを検索（前年データなど）"""
    files = glob.glob(os.path.join(BASE_DIR, f"*{keyword}*.csv"))
    if not files:
        return None
    result = max(files, key=os.path.getmtime)
    log(f"前年データ検出: {os.path.basename(result)}")
    return result

def load_csv(path):
    log(f"読み込み中: {os.path.basename(path)}")
    for enc in ["cp932", "utf-8-sig", "utf-8"]:
        try:
            return pd.read_csv(path, encoding=enc, low_memory=False)
        except Exception:
            continue
    raise ValueError(f"読み込み失敗: {path}")

def auto_fit_columns(ws, max_width=40):
    """列幅を自動調整する（最大40文字）"""
    for col_cells in ws.columns:
        length = 0
        for cell in col_cells:
            try:
                val_len = len(str(cell.value)) if cell.value else 0
                length = max(length, val_len)
            except Exception:
                pass
        col_letter = col_cells[0].column_letter
        ws.column_dimensions[col_letter].width = min(length + 2, max_width)

def get_season(month):
    """月からシーズンを返す（SS: 3-8月、AW: 9-2月）"""
    return "SS" if 3 <= month <= 8 else "AW"

def get_price_band(price):
    """価格帯を返す"""
    if price <= 5000:   return "① ~5,000円"
    if price <= 15000:  return "② 5,001~15,000円"
    if price <= 30000:  return "③ 15,001~30,000円"
    return "④ 30,001円~"

# =====================================
# データ読み込み・前処理
# =====================================
def load_and_merge_data():
    sales_file  = find_latest_file("*売上*.csv")
    stock_file  = find_latest_file("*在庫*.csv")
    master_file = find_latest_file("*マスタ*.csv")

    df_sales  = load_csv(sales_file)
    df_stock  = load_csv(stock_file)
    df_master = load_csv(master_file)

    # カラム名クリーンアップ
    df_sales.columns  = [c.strip() for c in df_sales.columns]
    df_stock.columns  = [c.strip() for c in df_stock.columns]
    df_master.columns = [c.strip() for c in df_master.columns]

    log(f"読み込み完了: 売上 {len(df_sales)}件 / 在庫 {len(df_stock)}件 / マスタ {len(df_master)}件")

    # --- カラムマッピング ---
    sc = df_sales.columns.tolist()
    mc = df_master.columns.tolist()
    ic = df_stock.columns.tolist()

    col_s_reg    = sc[0]           # A列: レシート番号
    col_s_store  = sc[2]           # C列: 店舗名
    col_s_key    = sc[3]           # D列: 商品コード（売上側）
    col_s_size   = sc[9]  if len(sc) > 9  else None
    col_s_price  = sc[14] if len(sc) > 14 else None  # O列: 税抜金額
    col_s_inbound= sc[20] if len(sc) > 20 else None  # U列: 客層区分
    col_s_age    = sc[21] if len(sc) > 21 else None  # V列: 拡張区分
    col_s_member = sc[24] if len(sc) > 24 else None  # Y列: 会員番号

    col_m_key    = mc[0]           # A列: 商品コード（マスタ側）
    col_m_name   = mc[2]           # C列: 商品名
    col_m_brand  = mc[6]           # G列: Brand
    col_m_cost   = mc[11]          # L列: 標準原価
    col_m_cat_l  = mc[12]          # M列: 商品種別1（大分類）
    col_m_cat_m  = mc[13]          # N列: 商品種別2（中分類）← Kankenはここ
    col_m_cat_s  = mc[14]          # O列: 商品種別3（小分類）
    col_m_color  = mc[20]          # U列: ColorName

    col_i_key    = ic[2]           # C列: 商品コード（在庫側）

    # 日付カラム検索
    col_date = None
    for c in sc:
        if any(x in c for x in ["日", "日付", "年月日", "Date"]):
            col_date = c
            break

    # 型変換
    df_sales[col_s_key]   = df_sales[col_s_key].astype(str)
    df_stock[col_i_key]   = df_stock[col_i_key].astype(str)
    df_master[col_m_key]  = df_master[col_m_key].astype(str)

    # 価格・原価のクリーンアップ
    for df, col in [(df_sales, col_s_price), (df_master, col_m_cost)]:
        if col:
            df[col] = (df[col].astype(str)
                       .str.replace("¥", "").str.replace(",", "").str.strip())
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # 在庫数カラム特定
    stock_val_col = ic[-1]
    for c in ic:
        if any(x in c for x in ["在庫", "数量", "残数", "現在"]):
            stock_val_col = c
            break

    # 受入数量カラム（消化率計算用）
    stock_recv_col = None
    for c in ic:
        if any(x in c for x in ["受入", "入荷", "仕入"]):
            stock_recv_col = c
            break

    # 売上 + マスタ 結合
    df_merged = pd.merge(df_sales, df_master,
                         left_on=col_s_key, right_on=col_m_key,
                         how="left", suffixes=("", "_master"))

    # 在庫集計（品番レベル）
    df_stock_sum = df_stock.groupby(col_i_key)[stock_val_col].sum().reset_index()
    df_merged = pd.merge(df_merged, df_stock_sum,
                         left_on=col_s_key, right_on=col_i_key, how="left")
    df_merged[stock_val_col] = df_merged[stock_val_col].fillna(0)

    # 受入数量も品番レベルで集計（消化率用）
    recv_by_item = None
    if stock_recv_col:
        recv_by_item = df_stock.groupby(col_i_key)[stock_recv_col].sum().reset_index()

    # 日付変換
    if col_date:
        df_merged[col_date] = pd.to_datetime(df_merged[col_date], errors="coerce")

    cols = {
        "s_reg": col_s_reg, "s_store": col_s_store, "s_key": col_s_key,
        "s_size": col_s_size, "s_price": col_s_price, "s_inbound": col_s_inbound,
        "s_age": col_s_age, "s_member": col_s_member, "date": col_date,
        "m_key": col_m_key, "m_name": col_m_name, "m_brand": col_m_brand,
        "m_cost": col_m_cost, "m_cat_l": col_m_cat_l, "m_cat_m": col_m_cat_m,
        "m_cat_s": col_m_cat_s, "m_color": col_m_color,
        "stock_val": stock_val_col, "stock_recv": stock_recv_col,
    }

    return df_merged, df_stock, df_master, df_sales, recv_by_item, cols


# =====================================
# 分析計算
# =====================================
def calc_wos(df_merged, cols):
    """WOS（在庫週数）を直近14日の平均売上で計算"""
    col_date = cols["date"]
    col_key  = cols["s_key"]
    col_reg  = cols["s_reg"]
    col_stk  = cols["stock_val"]
    if col_date and df_merged[col_date].notna().any():
        max_date     = df_merged[col_date].max()
        two_weeks_ago = max_date - timedelta(days=14)
        df_2w = df_merged[df_merged[col_date] >= two_weeks_ago]
        sales_2w = df_2w.groupby(col_key)[col_reg].count().reset_index(name="sales_2w")
        sales_2w["weekly_avg"] = sales_2w["sales_2w"] / 2.0
        df_merged = pd.merge(df_merged, sales_2w, on=col_key, how="left")
        df_merged["weekly_avg"]   = df_merged["weekly_avg"].fillna(0.01)
        df_merged["WOS(在庫週数)"] = (df_merged[col_stk] / df_merged["weekly_avg"]).round(1)
        return df_merged, max_date
    df_merged["weekly_avg"]   = 0.01
    df_merged["WOS(在庫週数)"] = "N/A"
    return df_merged, None


def calc_abc_xyz(df_merged, cols, max_date):
    """ABC/XYZ分析。中分類（商品種別2）も出力に含める"""
    col_key   = cols["s_key"]
    col_name  = cols["m_name"]
    col_cat_l = cols["m_cat_l"]
    col_cat_m = cols["m_cat_m"]  # N列 商品種別2
    col_price = cols["s_price"]
    col_cost  = cols["m_cost"]
    col_reg   = cols["s_reg"]
    col_stk   = cols["stock_val"]
    col_date  = cols["date"]

    # カラム存在確認（suffix対応）
    for attr in ["col_cat_l", "col_cat_m", "col_name", "col_cost"]:
        val = locals()[attr]
        if val and val not in df_merged.columns:
            alt = f"{val}_master"
            if alt in df_merged.columns:
                locals()[attr] = alt

    agg_dict = {
        "販売数":   (col_reg, "count"),
        "売上金額":  (col_price, "sum") if col_price else (col_reg, lambda x: 0),
        "原価合計":  (col_cost, "sum")  if col_cost  else (col_reg, lambda x: 0),
        "平均在庫数": (col_stk, "mean"),
    }
    abc_df = df_merged.groupby(
        [col_key, col_name, col_cat_l, col_cat_m]
    ).agg(**agg_dict).reset_index()

    abc_df["売価"]   = (abc_df["売上金額"] / abc_df["販売数"]).fillna(0).round(0)
    if col_cost:
        abc_df["粗利額"] = abc_df["売上金額"] - abc_df["原価合計"]
        abc_df["原価率"] = (abc_df["原価合計"] / abc_df["売上金額"].replace(0, 1)).fillna(0)
    else:
        abc_df["粗利額"] = 0
        abc_df["原価率"] = 0
    abc_df["回転率"] = (abc_df["販売数"] / abc_df["平均在庫数"].replace(0, 0.01)).round(2)

    abc_df = abc_df.sort_values("売上金額", ascending=False)
    total  = abc_df["売上金額"].sum()
    abc_df["売上構成"]  = (abc_df["売上金額"] / total).fillna(0)
    abc_df["累計売上構成"] = abc_df["売上構成"].cumsum()
    abc_df["ABCランク"] = abc_df["累計売上構成"].apply(
        lambda r: "A" if r <= 0.80 else ("B" if r <= 0.90 else "C")
    )

    # XYZ
    if col_date and max_date is not None:
        pivot = df_merged.pivot_table(
            index=col_key,
            columns=pd.Grouper(key=col_date, freq="D"),
            values=col_reg, aggfunc="count", fill_value=0
        )
        xyz = pd.DataFrame()
        xyz["mean"] = pivot.mean(axis=1)
        xyz["std"]  = pivot.std(axis=1)
        xyz["需要変動係数"] = (xyz["std"] / xyz["mean"].replace(0, 0.01)).fillna(0).round(2)
        xyz["XYZランク"]  = xyz["需要変動係数"].apply(
            lambda cv: "X" if cv <= 0.3 else ("Y" if cv <= 0.8 else "Z")
        )
        abc_df = pd.merge(abc_df, xyz[["需要変動係数", "XYZランク"]], on=col_key, how="left")
    else:
        abc_df["需要変動係数"] = 0
        abc_df["XYZランク"] = "-"

    abc_df["ABC×XYZ"] = abc_df["ABCランク"] + abc_df["XYZランク"]

    # 出力カラム整理
    output = abc_df[[
        col_key, col_name, col_cat_l, col_cat_m,
        "売価", "販売数", "売上金額", "原価率", "粗利額",
        "平均在庫数", "回転率", "売上構成", "累計売上構成",
        "ABCランク", "需要変動係数", "XYZランク", "ABC×XYZ"
    ]].rename(columns={
        col_key: "商品コード", col_name: "商品名",
        col_cat_l: "大分類",   col_cat_m: "中分類"
    })
    return output, abc_df


def calc_sell_through(df_merged, recv_by_item, cols):
    """消化率（販売数 / 受入数量）を品番レベルで計算"""
    col_key   = cols["s_key"]
    col_reg   = cols["s_reg"]
    col_cat_m = cols["m_cat_m"]
    col_stk   = cols["stock_val"]
    col_recv  = cols["stock_recv"]
    col_date  = cols["date"]

    base = df_merged.groupby(col_key).agg(
        販売数=(col_reg, "count"),
        売上金額=(cols["s_price"], "sum") if cols["s_price"] else (col_reg, lambda x: 0),
        現在庫=(col_stk, "mean"),
    ).reset_index()

    # 中分類付加
    cat_map = df_merged[[col_key, col_cat_m]].drop_duplicates(subset=[col_key])
    base = pd.merge(base, cat_map, on=col_key, how="left")

    # シーズン（売上日の月モード）
    if col_date:
        month_mode = (df_merged.dropna(subset=[col_date])
                      .groupby(col_key)[col_date]
                      .apply(lambda s: s.dt.month.mode()[0] if len(s) > 0 else 0))
        base["シーズン"] = month_mode.map(lambda m: get_season(int(m)) if m else "-")
    else:
        base["シーズン"] = "-"

    # 受入数量（消化率分母）
    if recv_by_item is not None:
        recv_col = recv_by_item.columns[-1]
        base = pd.merge(base, recv_by_item.rename(columns={recv_by_item.columns[0]: col_key}),
                        on=col_key, how="left")
        base[recv_col] = base[recv_col].fillna(0)
        base["消化率"] = (base["販売数"] / base[recv_col].replace(0, 1)).clip(0, 1).round(3)
        base = base.rename(columns={recv_col: "受入数量"})
    else:
        base["受入数量"] = 0
        base["消化率"]  = 0

    return base


def calc_kpi(df_merged, cols):
    """総合KPIを計算"""
    col_reg     = cols["s_reg"]
    col_price   = cols["s_price"]
    col_inbound = cols["s_inbound"]
    col_member  = cols["s_member"]

    total_sales    = len(df_merged)
    total_revenue  = df_merged[col_price].sum() if col_price else 0
    unique_tx      = df_merged[col_reg].nunique()
    avg_spend      = total_revenue / unique_tx if unique_tx > 0 else 0
    items_per_tx   = total_sales   / unique_tx if unique_tx > 0 else 0
    member_ratio   = df_merged[col_member].notna().mean() if col_member else 0
    inbound_ratio  = (df_merged[col_inbound].astype(str)
                      .str.contains("外国|インバウンド|INBOUND", case=False, na=False)
                      .mean() if col_inbound else 0)
    return {
        "total_sales": total_sales, "total_revenue": total_revenue,
        "unique_tx": unique_tx, "avg_spend": avg_spend,
        "items_per_tx": items_per_tx, "member_ratio": member_ratio,
        "inbound_ratio": inbound_ratio,
    }


# =====================================
# Excel出力
# =====================================
def write_summary_sheet(writer, kpi, cols, df_merged):
    summary_df = pd.DataFrame({
        "項目": [
            "全体の売上個数", "客数（取引件数）", "合計売上高（税抜）",
            "客単価", "セット率（点/客）", "インバウンド比率", "会員比率", "最終データ更新日"
        ],
        "値": [
            kpi["total_sales"],
            kpi["unique_tx"],
            f"¥{kpi['total_revenue']:,.0f}",
            f"¥{kpi['avg_spend']:,.0f}",
            f"{kpi['items_per_tx']:.2f}",
            f"{kpi['inbound_ratio']:.1%}",
            f"{kpi['member_ratio']:.1%}",
            datetime.now().strftime("%Y-%m-%d"),
        ]
    })
    summary_df.to_excel(writer, sheet_name="総合サマリー", index=False)
    ws = writer.sheets["総合サマリー"]
    auto_fit_columns(ws)


def write_category_sheet(writer, df_merged, cols):
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList

    col_reg   = cols["s_reg"]
    col_key   = cols["s_key"]
    col_cat_m = cols["m_cat_m"]
    col_stk   = cols["stock_val"]

    cat_summary = df_merged.groupby(col_cat_m).agg({col_reg: "count"}).rename(columns={col_reg: "売上数"})
    cat_summary["売上構成比"] = cat_summary["売上数"] / cat_summary["売上数"].sum()
    cat_stock = df_merged.drop_duplicates(subset=[col_key]).groupby(col_cat_m)[col_stk].sum()
    cat_summary = pd.merge(cat_summary, cat_stock, on=col_cat_m, how="left").rename(columns={col_stk: "在庫数"})
    cat_summary["在庫構成比"] = cat_summary["在庫数"] / cat_summary["在庫数"].sum()
    cat_summary.to_excel(writer, sheet_name="カテゴリー構成比")

    ws = writer.sheets["カテゴリー構成比"]
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row: cell.number_format = "0%"
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row: cell.number_format = "0%"

    chart = BarChart()
    chart.type = "bar"; chart.style = 11
    chart.title = "中分類別 売上vs在庫 バランス（構成比）"
    chart.dLbls = DataLabelList(); chart.dLbls.showVal = True; chart.dLbls.numFmt = "0%"
    sales_ref = Reference(ws, min_col=3, min_row=1, max_row=len(cat_summary)+1)
    chart.add_data(sales_ref, titles_from_data=True)
    stock_ref = Reference(ws, min_col=5, min_row=1, max_row=len(cat_summary)+1)
    chart.add_data(stock_ref, titles_from_data=True)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(cat_summary)+1)
    chart.set_categories(cats)
    chart.height = 20; chart.width = 20
    ws.add_chart(chart, "G2")
    auto_fit_columns(ws)


def write_trend_sheet(writer, df_merged, cols):
    from openpyxl.formatting.rule import ColorScaleRule

    col_date  = cols["date"]
    col_cat_m = cols["m_cat_m"]
    if not col_date or not df_merged[col_date].notna().any():
        return
    max_date    = df_merged[col_date].max()
    one_month   = max_date - timedelta(days=30)
    df_1m       = df_merged[df_merged[col_date] >= one_month]
    trend_1m    = df_1m.groupby([pd.Grouper(key=col_date, freq="D"), col_cat_m]).size().unstack().fillna(0)
    trend_1m.to_excel(writer, sheet_name="直近1ヶ月トレンド")
    ws = writer.sheets["直近1ヶ月トレンド"]
    rule = ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="FF6347")
    ws.conditional_formatting.add("B2:AZ40", rule)


def write_abc_xyz_sheet(writer, abc_xyz_output):
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import CellIsRule, DataBarRule

    abc_xyz_output.to_excel(writer, sheet_name="ABC_XYZ分析", index=False, startrow=11)
    ws = writer.sheets["ABC_XYZ分析"]

    # 凡例
    legends = [
        ("A1", "【ABC×XYZランクの見方】"),
        ("A2", "AX: 超主力・安定（欠品厳禁・最優先確保）"),
        ("A3", "AY: 主力・中変動（在庫維持・安定供給）"),
        ("A4", "AZ: 主力・不安定（波あり・機動的発注）"),
        ("A5", "BX: 準主力・安定（定番品・在庫効率重視）"),
        ("A6", "BY: 準主力・中変動（動向注視・適正在庫）"),
        ("A7", "BZ: 準主力・不安定（販促または処分検討）"),
        ("A8", "CX: 下位・安定（ロングテール・低在庫維持）"),
        ("A9", "CY: 下位・中変動（需要がある限り維持）"),
        ("A10", "CZ: 下位・不安定（非効率・取扱終了の検討）"),
    ]
    for cell_addr, text in legends:
        ws[cell_addr] = text
    ws.freeze_panes = "A13"

    # 条件付き書式（R列=18列目がABC×XYZ）
    abc_xyz_col_idx = abc_xyz_output.columns.tolist().index("ABC×XYZ") + 1
    abc_xyz_col_letter = ws.cell(row=12, column=abc_xyz_col_idx).column_letter
    abc_xyz_range = f"{abc_xyz_col_letter}13:{abc_xyz_col_letter}3000"

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add(abc_xyz_range, CellIsRule(operator="equal", formula=['"AX"'], fill=green_fill))
    ws.conditional_formatting.add(abc_xyz_range, CellIsRule(operator="equal", formula=['"CZ"'], fill=red_fill))

    # 回転率 データバー（K列=11列目）
    turn_col_idx    = abc_xyz_output.columns.tolist().index("回転率") + 1
    turn_col_letter = ws.cell(row=12, column=turn_col_idx).column_letter
    ws.conditional_formatting.add(
        f"{turn_col_letter}13:{turn_col_letter}3000",
        DataBarRule(start_type="min", end_type="max", color="638EC6", showValue=True)
    )

    # ％表示
    for col_name_str in ["原価率", "売上構成", "累計売上構成"]:
        if col_name_str in abc_xyz_output.columns:
            ci = abc_xyz_output.columns.tolist().index(col_name_str) + 1
            cl = ws.cell(row=12, column=ci).column_letter
            for row in ws.iter_rows(min_row=13, min_col=ci, max_col=ci):
                for cell in row: cell.number_format = "0%"

    auto_fit_columns(ws)


def write_detail_sheet(writer, df_merged, abc_df_raw, cols):
    """詳細シート: WOS昇順 → ABCランク → 売上金額降順でソート"""
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import CellIsRule

    col_date  = cols["date"]
    col_store = cols["s_store"]
    col_key   = cols["s_key"]
    col_name  = cols["m_name"]
    col_color = cols["m_color"]
    col_size  = cols["s_size"]
    col_cat_l = cols["m_cat_l"]
    col_cat_m = cols["m_cat_m"]
    col_price = cols["s_price"]
    col_stk   = cols["stock_val"]
    col_reg   = cols["s_reg"]

    # ABCランクをマージ
    abc_key_col = cols["s_key"]
    abc_rank_map = abc_df_raw.set_index(abc_key_col)[["ABCランク", "売上金額"]].to_dict("index") \
        if "ABCランク" in abc_df_raw.columns else {}

    detail_cols = [c for c in [col_date, col_store, col_key, col_name, col_color,
                                col_size, col_cat_l, col_cat_m, col_price, col_stk, "WOS(在庫週数)"]
                   if c and c in df_merged.columns]

    df_detail = df_merged[detail_cols].drop_duplicates().copy()

    # ABCランク付加
    if abc_rank_map:
        df_detail["ABCランク"] = df_detail[col_key].map(
            lambda k: abc_rank_map.get(k, {}).get("ABCランク", "C")
        )
        df_detail["_売上金額"] = df_detail[col_key].map(
            lambda k: abc_rank_map.get(k, {}).get("売上金額", 0)
        )
        rank_order = {"A": 0, "B": 1, "C": 2}
        df_detail["_ランク順"] = df_detail["ABCランク"].map(lambda r: rank_order.get(r, 9))
    else:
        df_detail["ABCランク"] = "-"
        df_detail["_売上金額"] = 0
        df_detail["_ランク順"] = 9

    # WOS数値化（ソート用）
    df_detail["_wos_num"] = pd.to_numeric(df_detail["WOS(在庫週数)"], errors="coerce").fillna(999)

    # ソート: WOS昇順 → ABCランク → 売上金額降順
    df_detail = df_detail.sort_values(
        ["_wos_num", "_ランク順", "_売上金額"],
        ascending=[True, True, False]
    ).drop(columns=["_wos_num", "_ランク順", "_売上金額"])

    output_cols = detail_cols + ["ABCランク"]
    df_detail[output_cols].head(5000).to_excel(writer, sheet_name="売上在庫詳細(WOS昇順)", index=False)

    ws = writer.sheets["売上在庫詳細(WOS昇順)"]
    # WOS列を動的に特定
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    wos_col_idx = next((i+1 for i, h in enumerate(header) if h == "WOS(在庫週数)"), None)
    if wos_col_idx:
        wos_letter = ws.cell(row=1, column=wos_col_idx).column_letter
        red_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        wos_range   = f"{wos_letter}2:{wos_letter}5001"
        ws.conditional_formatting.add(wos_range, CellIsRule(operator="lessThan",    formula=["2"],  fill=red_fill))
        ws.conditional_formatting.add(wos_range, CellIsRule(operator="greaterThan", formula=["20"], fill=yellow_fill))

    auto_fit_columns(ws)


def write_store_comparison_sheet(writer, df_merged, cols):
    """店舗別比較シート: 店舗×中分類 の売上数・売上金額・WOS"""
    col_store = cols["s_store"]
    col_cat_m = cols["m_cat_m"]
    col_reg   = cols["s_reg"]
    col_price = cols["s_price"]

    # 売上数ピボット
    pivot_cnt = df_merged.pivot_table(
        index=col_store, columns=col_cat_m, values=col_reg,
        aggfunc="count", fill_value=0
    )
    pivot_cnt.columns = [f"{c}_販売数" for c in pivot_cnt.columns]

    # 売上金額ピボット
    if col_price:
        pivot_amt = df_merged.pivot_table(
            index=col_store, columns=col_cat_m, values=col_price,
            aggfunc="sum", fill_value=0
        )
        pivot_amt.columns = [f"{c}_売上金額" for c in pivot_amt.columns]
        result = pd.concat([pivot_cnt, pivot_amt], axis=1).sort_index(axis=1)
    else:
        result = pivot_cnt

    result.to_excel(writer, sheet_name="店舗別比較")
    ws = writer.sheets["店舗別比較"]
    auto_fit_columns(ws, max_width=20)


def write_sell_through_sheet(writer, sell_through_df):
    """消化率・シーズン分析シート"""
    from openpyxl.formatting.rule import ColorScaleRule

    sell_through_df.to_excel(writer, sheet_name="消化率・シーズン分析", index=False)
    ws = writer.sheets["消化率・シーズン分析"]

    # 消化率列を特定してカラースケール適用
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    st_col_idx = next((i+1 for i, h in enumerate(header) if h == "消化率"), None)
    if st_col_idx:
        st_letter = ws.cell(row=1, column=st_col_idx).column_letter
        for row in ws.iter_rows(min_row=2, min_col=st_col_idx, max_col=st_col_idx):
            for cell in row: cell.number_format = "0%"
        rule = ColorScaleRule(
            start_type="num", start_value=0,   start_color="FFC7CE",
            mid_type="num",   mid_value=0.5,   mid_color="FFEB9C",
            end_type="num",   end_value=1,     end_color="C6EFCE"
        )
        ws.conditional_formatting.add(f"{st_letter}2:{st_letter}5000", rule)

    auto_fit_columns(ws)


def write_kanken_sheet(writer, df_merged, cols):
    """Kanken専用分析シート（中分類=商品種別2でフィルタ）"""
    col_cat_m = cols["m_cat_m"]
    col_store = cols["s_store"]
    col_color = cols["m_color"]
    col_size  = cols["s_size"]
    col_reg   = cols["s_reg"]

    df_kanken = df_merged[
        df_merged[col_cat_m].astype(str).str.contains("Kanken|カンケン|kanken", case=False, na=False)
    ]
    if df_kanken.empty:
        return

    # カラー別売上
    color_sales = df_kanken.groupby(col_color)[col_reg].count().sort_values(ascending=False).reset_index()
    color_sales.columns = ["カラー", "販売数"]

    # 店舗×サイズのクロス集計
    if col_size:
        cross = df_kanken.pivot_table(
            index=col_store, columns=col_size, values=col_reg,
            aggfunc="count", fill_value=0
        )
    else:
        cross = pd.DataFrame()



    color_sales.to_excel(writer, sheet_name="Kanken詳細", index=False)
    ws_k = writer.sheets["Kanken詳細"]
    auto_fit_columns(ws_k)

    if not cross.empty:
        start_row = len(color_sales) + 3
        ws_k.cell(row=start_row, column=1, value="【店舗×サイズ クロス集計】")
        cross.to_excel(writer, sheet_name="Kanken詳細",
                       startrow=start_row, startcol=0)


def write_inbound_sheet(writer, df_merged, cols):
    """インバウンド分析シート"""
    col_inbound = cols["s_inbound"]
    col_cat_m   = cols["m_cat_m"]
    col_price   = cols["s_price"]
    col_reg     = cols["s_reg"]

    if not col_inbound:
        return

    df_merged["_inbound_flag"] = df_merged[col_inbound].astype(str).str.contains(
        "外国|インバウンド|INBOUND", case=False, na=False
    )
    inbound_cat = df_merged[df_merged["_inbound_flag"]].groupby(col_cat_m).agg(
        販売数=(col_reg, "count"),
        売上金額=(col_price, "sum") if col_price else (col_reg, lambda x: 0),
    ).sort_values("販売数", ascending=False)

    inbound_cat["構成比"] = inbound_cat["販売数"] / inbound_cat["販売数"].sum()
    inbound_cat.to_excel(writer, sheet_name="インバウンド分析")
    ws = writer.sheets["インバウンド分析"]
    auto_fit_columns(ws)


# =====================================
# DB保存（正規化版）
# =====================================
def save_to_db(df_sales, df_stock, df_master, cols):
    """
    正規化DB設計:
      sales_raw    : 売上トランザクション（マスタ未結合）
      stock_history: 在庫スナップショット（重複防止付き）
      master_data  : 商品マスタ（最新を上書き）
    """
    log("データベースを更新中（v2・正規化版）...")
    conn = sqlite3.connect(DB_FILE)

    # --- master_data: 最新を上書き ---
    df_master.to_sql("master_data", conn, if_exists="replace", index=False)

    # --- sales_raw: 日付範囲で重複防止 ---
    col_date = cols["date"]
    if col_date and df_sales[col_date].notna().any():
        # 元のCSVのdf_salesはdatetime変換前なので文字列のまま保存
        df_sales_save = df_sales.copy()
        df_sales_save[col_date] = df_sales_save[col_date].astype(str)
        min_d = df_sales_save[col_date].min()
        max_d = df_sales_save[col_date].max()
        try:
            conn.execute(f'DELETE FROM sales_raw WHERE "{col_date}" BETWEEN ? AND ?', (min_d, max_d))
            conn.commit()
        except sqlite3.OperationalError:
            pass
        df_sales_save.to_sql("sales_raw", conn, if_exists="append", index=False)
    else:
        df_sales.to_sql("sales_raw", conn, if_exists="append", index=False)

    # --- stock_history: 同日データ重複防止 ---
    today_str = datetime.now().strftime("%Y-%m-%d")
    try:
        conn.execute("DELETE FROM stock_history WHERE 実行日 = ?", (today_str,))
        conn.commit()
    except sqlite3.OperationalError:
        pass
    df_stock_save = df_stock.copy()
    df_stock_save["実行日"] = today_str
    df_stock_save.to_sql("stock_history", conn, if_exists="append", index=False)

    conn.close()
    log("DB更新完了。")


def save_prev_year_to_db(df_prev):
    """前年売上データをsales_raw_prevとしてDB保存（重複防止付き）"""
    conn = sqlite3.connect(DB_FILE)
    # 前年データは全期間を一括置換
    df_prev_save = df_prev.copy()
    # 日付列を文字列化
    date_candidates = [c for c in df_prev.columns if any(x in c for x in ["日","日付","年月日","Date"])]
    if date_candidates:
        df_prev_save[date_candidates[0]] = df_prev_save[date_candidates[0]].astype(str)
    df_prev_save.to_sql("sales_raw_prev", conn, if_exists="replace", index=False)
    conn.close()
    log(f"前年データをDBに保存しました（{len(df_prev)}件）。")


# =====================================
# 前年同月比計算・Excel出力
# =====================================
def calc_yoy(df_current, df_prev, cols):
    """
    前年同月比を計算する。
    - 今年データ: df_current（売上済みデータ＋マスタ結合済み）
    - 前年データ: df_prev（生のCSV）
    - 集計軸: 月別 × 中分類（商品種別2）
    """
    col_date  = cols["date"]
    col_cat_m = cols["m_cat_m"]
    col_reg   = cols["s_reg"]
    col_price = cols["s_price"]

    if not col_date:
        return None

    # 今年: 月別×中分類 売上数・売上金額
    df_cur = df_current.copy()
    df_cur["年月"] = df_cur[col_date].dt.to_period("M").astype(str)
    df_cur["年"]  = df_cur[col_date].dt.year
    cur_agg = df_cur.groupby(["年月", col_cat_m]).agg(
        今年販売数=(col_reg, "count"),
        今年売上金額=(col_price, "sum") if col_price else (col_reg, lambda x: 0),
    ).reset_index().rename(columns={col_cat_m: "中分類"})

    # 前年データの日付・売上列を特定
    prev_cols = df_prev.columns.tolist()
    prev_date_col  = next((c for c in prev_cols if any(x in c for x in ["日","日付","年月日","Date"])), None)
    prev_price_col = prev_cols[14] if len(prev_cols) > 14 else None
    prev_reg_col   = prev_cols[0]
    # 前年の中分類はD列(index3)の商品コードをキーにマスタと結合できないため
    # 代わりに商品名列(index4付近)から代用、または売上CSVに含まれる分類列を使う
    # ここでは「表記部門名」(index4)を前年中分類の代替として使用
    prev_cat_col = prev_cols[4] if len(prev_cols) > 4 else None

    if not prev_date_col:
        log("警告: 前年データに日付列が見つかりません。前年比シートをスキップします。")
        return None

    df_py = df_prev.copy()
    df_py[prev_date_col] = pd.to_datetime(df_py[prev_date_col], errors="coerce")
    if prev_price_col:
        df_py[prev_price_col] = (df_py[prev_price_col].astype(str)
                                 .str.replace("¥","").str.replace(",","").str.strip())
        df_py[prev_price_col] = pd.to_numeric(df_py[prev_price_col], errors="coerce").fillna(0)

    df_py["年月_前年"] = df_py[prev_date_col].dt.to_period("M").astype(str)
    # 前年月を今年月に合わせるため +1年
    df_py["年月"] = df_py[prev_date_col].apply(
        lambda d: f"{d.year+1}-{d.month:02d}" if pd.notna(d) else None
    )

    if prev_cat_col:
        prev_agg = df_py.groupby(["年月", prev_cat_col]).agg(
            前年販売数=(prev_reg_col, "count"),
            前年売上金額=(prev_price_col, "sum") if prev_price_col else (prev_reg_col, lambda x: 0),
        ).reset_index().rename(columns={prev_cat_col: "中分類"})
    else:
        prev_agg = df_py.groupby(["年月"]).agg(
            前年販売数=(prev_reg_col, "count"),
            前年売上金額=(prev_price_col, "sum") if prev_price_col else (prev_reg_col, lambda x: 0),
        ).reset_index()
        prev_agg["中分類"] = "全体"

    # 結合
    yoy = pd.merge(cur_agg, prev_agg, on=["年月", "中分類"], how="outer").fillna(0)
    yoy = yoy.sort_values(["年月", "中分類"]).reset_index(drop=True)

    # 前年比計算（0除算を防ぐ）
    yoy["販売数前年比"]   = yoy.apply(
        lambda r: r["今年販売数"] / r["前年販売数"] if r["前年販売数"] > 0 else None, axis=1
    )
    yoy["売上金額前年比"] = yoy.apply(
        lambda r: r["今年売上金額"] / r["前年売上金額"] if r["前年売上金額"] > 0 else None, axis=1
    )
    return yoy


def write_yoy_sheet(writer, yoy_df):
    """前年同月比シートを出力（カラースケール付き）"""
    from openpyxl.formatting.rule import ColorScaleRule

    if yoy_df is None or yoy_df.empty:
        return

    yoy_df.to_excel(writer, sheet_name="前年同月比", index=False)
    ws = writer.sheets["前年同月比"]

    # 前年比列をヘッダーから動的取得してカラースケール適用
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for col_name in ["販売数前年比", "売上金額前年比"]:
        if col_name in header:
            ci = header.index(col_name) + 1
            cl = ws.cell(row=1, column=ci).column_letter
            for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
                for cell in row:
                    cell.number_format = "0%"
            rule = ColorScaleRule(
                start_type="num", start_value=0.7, start_color="FFC7CE",   # 70%未満=赤
                mid_type="num",   mid_value=1.0,   mid_color="FFFFFF",    # 100%=白
                end_type="num",   end_value=1.3,   end_color="C6EFCE"     # 130%超=緑
            )
            ws.conditional_formatting.add(f"{cl}2:{cl}5000", rule)

    auto_fit_columns(ws)


# =====================================
# メイン処理
# =====================================
def main():
    try:
        os.makedirs(V2_DIR, exist_ok=True)
        if os.path.exists(LOG_FILE):
            os.remove(LOG_FILE)
        log("=== FJALLRAVEN MD分析ツール v2 開始 ===")

        try:
            import openpyxl
        except ImportError:
            log("エラー: openpyxlが見つかりません。pip install openpyxl を実行してください。")
            return

        # 1. データ読み込み・前処理
        df_merged, df_stock, df_master, df_sales, recv_by_item, cols = load_and_merge_data()

        # 2. 前年データ読み込み（ファイル名に「2025」を含むCSV）
        prev_year_file = find_file_by_keyword("2025")
        df_prev = None
        if prev_year_file:
            df_prev = load_csv(prev_year_file)
            df_prev.columns = [c.strip() for c in df_prev.columns]
            log(f"前年データ読み込み完了: {len(df_prev)}件")
        else:
            log("前年データが見つかりませんでした（前年比シートはスキップ）。")

        # 3. 分析計算
        df_merged, max_date = calc_wos(df_merged, cols)
        abc_xyz_output, abc_df_raw = calc_abc_xyz(df_merged, cols, max_date)
        sell_through_df = calc_sell_through(df_merged, recv_by_item, cols)
        kpi = calc_kpi(df_merged, cols)
        yoy_df = calc_yoy(df_merged, df_prev, cols) if df_prev is not None else None

        # 4. DB保存
        save_to_db(df_sales, df_stock, df_master, cols)
        if df_prev is not None:
            save_prev_year_to_db(df_prev)

        # 5. Excel出力
        log("Excelレポートを生成中...")
        try:
            with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
                write_summary_sheet(writer, kpi, cols, df_merged)
                write_category_sheet(writer, df_merged, cols)
                write_trend_sheet(writer, df_merged, cols)
                write_abc_xyz_sheet(writer, abc_xyz_output)
                write_detail_sheet(writer, df_merged, abc_df_raw, cols)
                write_store_comparison_sheet(writer, df_merged, cols)
                write_sell_through_sheet(writer, sell_through_df)
                write_kanken_sheet(writer, df_merged, cols)
                write_inbound_sheet(writer, df_merged, cols)
                write_yoy_sheet(writer, yoy_df)

        except PermissionError:
            log("【エラー】Excelファイルが開いています。閉じてから再実行してください。")
            return

        log(f"=== 完了: {os.path.basename(OUTPUT_EXCEL)} ===")

    except Exception as e:
        import traceback
        log(f"致命的なエラー: {e}")
        log(traceback.format_exc())


if __name__ == "__main__":
    main()
