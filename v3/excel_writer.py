"""
Excel出力モジュール
各シートの書き出しと装飾（条件付き書式、グラフ等）を担当する。
"""
import pandas as pd
from datetime import datetime, timedelta
from utils import get_config, auto_fit_columns, log


def write_summary_sheet(writer, kpi, cols, df_merged):
    summary_df = pd.DataFrame({
        "項目": [
            "全体の売上個数", "客数（取引件数）", "合計売上高（税抜）",
            "客単価", "セット率（点/客）", "インバウンド比率", "会員比率", "最終データ更新日"
        ],
        "値": [
            kpi["total_sales"], kpi["unique_tx"],
            f"¥{kpi['total_revenue']:,.0f}", f"¥{kpi['avg_spend']:,.0f}",
            f"{kpi['items_per_tx']:.2f}", f"{kpi['inbound_ratio']:.1%}",
            f"{kpi['member_ratio']:.1%}", datetime.now().strftime("%Y-%m-%d"),
        ]
    })
    summary_df.to_excel(writer, sheet_name="総合サマリー", index=False)
    ws = writer.sheets["総合サマリー"]
    
    from openpyxl.styles import PatternFill, Font
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100")
    
    if kpi.get("inbound_ratio", 0) >= 0.3:
        ws["B7"].fill = green_fill
        ws["B7"].font = green_font
    if kpi.get("member_ratio", 0) >= 0.3:
        ws["B8"].fill = green_fill
        ws["B8"].font = green_font
        
    auto_fit_columns(ws)


def write_category_sheet(writer, df_merged, cols):
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList

    col_reg = cols["s_reg"]; col_key = cols["s_key"]
    col_cat_m = cols["m_cat_m"]; col_stk = cols["stock_val"]

    cat = df_merged.groupby(col_cat_m).agg({col_reg: "count"}).rename(columns={col_reg: "売上数"})
    cat["売上構成比"] = cat["売上数"] / cat["売上数"].sum()
    cat_stock = df_merged.drop_duplicates(subset=[col_key]).groupby(col_cat_m)[col_stk].sum()
    cat = pd.merge(cat, cat_stock, on=col_cat_m, how="left").rename(columns={col_stk: "在庫数"})
    cat["在庫構成比"] = cat["在庫数"] / cat["在庫数"].sum()
    
    # 売上数の多い順にソート（円グラフで右側から大きい順に表示するため）
    cat = cat.sort_values("売上数", ascending=False)
    
    cat.to_excel(writer, sheet_name="カテゴリー構成比")

    ws = writer.sheets["カテゴリー構成比"]
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row: cell.number_format = "0%"
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row: cell.number_format = "0%"

    chart = BarChart(); chart.type = "bar"; chart.style = 11
    chart.title = "中分類別 売上vs在庫 バランス（構成比）"
    chart.dLbls = DataLabelList(); chart.dLbls.showVal = True; chart.dLbls.numFmt = "0%"
    chart.add_data(Reference(ws, min_col=3, min_row=1, max_row=len(cat)+1), titles_from_data=True)
    chart.add_data(Reference(ws, min_col=5, min_row=1, max_row=len(cat)+1), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=len(cat)+1))
    # 横棒グラフで上から大きい順（行の順）に表示させるため、カテゴリー軸(x_axis)を反転
    chart.x_axis.scaling.orientation = "maxMin"
    chart.height = 20; chart.width = 20
    ws.add_chart(chart, "G2")

    # 売上構成比の円グラフ
    from openpyxl.chart import PieChart
    pie = PieChart()
    pie.title = "売上構成比"
    pie.add_data(Reference(ws, min_col=3, min_row=1, max_row=len(cat)+1), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=1, min_row=2, max_row=len(cat)+1))
    pie.dLbls = DataLabelList()
    pie.dLbls.showPercent = True
    pie.dLbls.showCatName = True
    pie.dLbls.showVal = False
    pie.height = 18; pie.width = 18
    ws.add_chart(pie, "G22")

    auto_fit_columns(ws)


def write_trend_sheet(writer, df_merged, cols):
    from openpyxl.formatting.rule import ColorScaleRule
    col_date = cols["date"]; col_cat_m = cols["m_cat_m"]
    if not col_date or not df_merged[col_date].notna().any():
        return

    max_d = df_merged[col_date].max()
    df_1m = df_merged[df_merged[col_date] >= max_d - timedelta(days=30)]
    trend = df_1m.groupby([pd.Grouper(key=col_date, freq="D"), col_cat_m]).size().unstack().fillna(0)
    trend.to_excel(writer, sheet_name="直近1ヶ月トレンド")
    ws = writer.sheets["直近1ヶ月トレンド"]
    ws.conditional_formatting.add("B2:AZ40",
        ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="FF6347"))

    # 折れ線グラフを追加
    from openpyxl.chart import LineChart, Reference as ChartRef
    num_rows = len(trend)
    num_cols = len(trend.columns)
    if num_rows > 0 and num_cols > 0:
        line = LineChart()
        line.title = "直近1ヶ月 カテゴリー別売上トレンド"
        line.style = 10
        line.y_axis.title = "販売数"
        line.x_axis.title = "日付"
        data = ChartRef(ws, min_col=2, min_row=1, max_col=num_cols+1, max_row=num_rows+1)
        cats = ChartRef(ws, min_col=1, min_row=2, max_row=num_rows+1)
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.height = 18; line.width = 30
        # データ行の下に配置
        ws.add_chart(line, f"A{num_rows + 4}")


def write_abc_xyz_sheet(writer, abc_xyz_output):
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import CellIsRule, DataBarRule
    cfg = get_config()

    abc_xyz_output.to_excel(writer, sheet_name="ABC_XYZ分析", index=False, startrow=11)
    ws = writer.sheets["ABC_XYZ分析"]

    legends = [
        ("A1", "【ABC×XYZランクの見方】"),
        ("A2", "AX: 超主力・安定（欠品厳禁・最優先確保）"),
        ("A3", "AY: 主力・中変動（在庫維持・安定供給）"),
        ("A4", "AZ: 主力・不安定（波あり・機動的発注）"),
        ("A5", "BX: 準主力・安定（定番品・在庫効率重視）"),
        ("A6", "BY: 準主力・中変動（動向注視・適正在庫）"),
        ("A7", "BZ: 準主力・不安定（販促または処分検討）"),
        ("A8", "CX: 下位・安定（ロングテール・低在庫維持）"),
        ("A9", "CY: 下位・中変動（需要がある限り維持）"),
        ("A10", "CZ: 下位・不安定（非効率・取扱終了の検討）"),
    ]
    for addr, txt in legends: ws[addr] = txt
    ws.freeze_panes = "A13"

    cols_list = abc_xyz_output.columns.tolist()
    abc_col = cols_list.index("ABC×XYZ") + 1
    abc_letter = ws.cell(row=12, column=abc_col).column_letter
    rng = f"{abc_letter}13:{abc_letter}3000"
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"AX"'], fill=green))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"CZ"'], fill=red))

    turn_col = cols_list.index("回転率") + 1
    turn_letter = ws.cell(row=12, column=turn_col).column_letter
    ws.conditional_formatting.add(f"{turn_letter}13:{turn_letter}3000",
        DataBarRule(start_type="min", end_type="max", color="638EC6", showValue=True))

    for cn in ["原価率", "売上構成", "累計売上構成"]:
        if cn in cols_list:
            ci = cols_list.index(cn) + 1
            for row in ws.iter_rows(min_row=13, min_col=ci, max_col=ci):
                for cell in row: cell.number_format = "0%"
    auto_fit_columns(ws)


def write_detail_sheet(writer, df_merged, abc_df_raw, sell_through_df, cols):
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import CellIsRule
    cfg = get_config()
    max_rows = cfg["analysis"]["detail_max_rows"]
    wos_red  = cfg["analysis"]["wos_red_below"]
    wos_yel  = cfg["analysis"]["wos_yellow_above"]

    col_key   = cols["s_key"]
    col_name  = cols["m_name"]
    col_cat_l = cols["m_cat_l"]
    col_cat_m = cols["m_cat_m"]
    col_color = cols["m_color"]
    col_price = cols["s_price"]
    col_stk   = cols["stock_val"]
    col_reg   = cols["s_reg"]

    # --- 品番レベルで集計（1品番1行） ---
    agg_dict = {"売上数": (col_reg, "count")}
    if col_price:
        agg_dict["売上金額"] = (col_price, "sum")
    if col_stk:
        agg_dict["在庫数"] = (col_stk, "mean")  # 品番ごとに同じ値なのでmeanで取得

    group_keys = [k for k in [col_key, col_name, col_cat_l, col_cat_m, col_color]
                  if k and k in df_merged.columns]
    df_d = df_merged.groupby(group_keys).agg(**agg_dict).reset_index()
    df_d["在庫数"] = df_d["在庫数"].round(0).astype(int)

    # WOS を品番単位でマージ
    if "WOS(在庫週数)" in df_merged.columns:
        wos_map = (df_merged[[col_key, "WOS(在庫週数)"]]
                   .drop_duplicates(subset=[col_key])
                   .set_index(col_key)["WOS(在庫週数)"])
        df_d["WOS(在庫週数)"] = df_d[col_key].map(wos_map)

    # テンポサービス（受入数量0）を除外
    if sell_through_df is not None and "受入数量" in sell_through_df.columns and sell_through_df["受入数量"].sum() > 0:
        valid_keys = sell_through_df[sell_through_df["受入数量"] > 0][col_key]
        df_d = df_d[df_d[col_key].isin(valid_keys)]

    # 消化率をマージ
    if sell_through_df is not None and "消化率" in sell_through_df.columns:
        st_map = sell_through_df.set_index(col_key)["消化率"]
        df_d["消化率"] = df_d[col_key].map(st_map).fillna(0)

    # ABCランク付与
    abc_map = abc_df_raw.drop_duplicates(subset=[col_key]).set_index(col_key)[["ABCランク", "売上金額"]].to_dict("index") \
        if "ABCランク" in abc_df_raw.columns else {}
    df_d["ABCランク"] = df_d[col_key].map(lambda k: abc_map.get(k, {}).get("ABCランク", "C"))

    # ソート：WOS昇順 → ABCランク順 → 売上金額降順
    sort_col = "売上金額" if "売上金額" in df_d.columns else col_key
    df_d["_ランク順"] = df_d["ABCランク"].map({"A": 0, "B": 1, "C": 2})
    df_d["_wos"]     = pd.to_numeric(df_d.get("WOS(在庫週数)", 999), errors="coerce").fillna(999)
    df_d = df_d.sort_values(["_wos", "_ランク順", sort_col], ascending=[True, True, False])
    df_d = df_d.drop(columns=["_wos", "_ランク順"])

    df_d.head(max_rows).to_excel(writer, sheet_name="売上在庫詳細(WOS昇順)", index=False)

    ws = writer.sheets["売上在庫詳細(WOS昇順)"]
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]

    # WOS 条件付き書式
    wos_i = next((i+1 for i, h in enumerate(header) if h == "WOS(在庫週数)"), None)
    if wos_i:
        wl = ws.cell(row=1, column=wos_i).column_letter
        red_f  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yel_f  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        ws.conditional_formatting.add(f"{wl}2:{wl}{max_rows+1}",
            CellIsRule(operator="lessThan", formula=[str(wos_red)], fill=red_f))
        ws.conditional_formatting.add(f"{wl}2:{wl}{max_rows+1}",
            CellIsRule(operator="greaterThan", formula=[str(wos_yel)], fill=yel_f))
    auto_fit_columns(ws)


def write_store_comparison_sheet(writer, df_merged, df_stock, cols):
    from openpyxl.formatting.rule import ColorScaleRule
    col_store=cols["s_store"]; col_cat_m=cols["m_cat_m"]
    col_reg=cols["s_reg"]
    
    # 心斎橋パルコと大丸心斎橋を統合する処理
    df_merged_temp = df_merged.copy()
    df_merged_temp[col_store] = df_merged_temp[col_store].replace({
        "FJALLRAVEN by 3NITY 心斎橋パルコ": "FJALLRAVEN by 3NITY 大丸心斎橋（心斎橋パルコ）",
        "FJALLRAVEN by 3NITY 大丸心斎橋": "FJALLRAVEN by 3NITY 大丸心斎橋（心斎橋パルコ）"
    })
    
    pc = df_merged_temp.pivot_table(index=col_store, columns=col_cat_m, values=col_reg, aggfunc="count", fill_value=0)
    
    # 在庫一覧から店舗コードを取得（0列目: 拠点コード, 1列目: 拠点名）
    col_stock_code = df_stock.columns[0]
    col_stock_store = df_stock.columns[1]
    
    df_stock_temp = df_stock.copy()
    df_stock_temp[col_stock_store] = df_stock_temp[col_stock_store].replace({
        "FJALLRAVEN by 3NITY 心斎橋パルコ": "FJALLRAVEN by 3NITY 大丸心斎橋（心斎橋パルコ）",
        "FJALLRAVEN by 3NITY 大丸心斎橋": "FJALLRAVEN by 3NITY 大丸心斎橋（心斎橋パルコ）"
    })
    store_map = df_stock_temp.drop_duplicates(subset=[col_stock_store]).set_index(col_stock_store)[col_stock_code]
    
    pc = pc.reset_index()
    pc.insert(0, "店舗コード", pc[col_store].map(store_map))
    pc.columns.name = None
    
    # 店舗コードのソート処理（数字→英数字の順）
    pc["_sort_num"] = pd.to_numeric(pc["店舗コード"], errors="coerce")
    pc = pc.sort_values(by=["_sort_num", "店舗コード"]).drop(columns=["_sort_num"])
    
    pc.to_excel(writer, sheet_name="店舗別比較", index=False)
    ws = writer.sheets["店舗別比較"]
    
    # C列以降（カテゴリー別売上）にヒートマップ（カラースケール）を適用
    max_col_letter = ws.cell(row=1, column=ws.max_column).column_letter
    rule = ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="FF9999")
    ws.conditional_formatting.add(f"C2:{max_col_letter}{ws.max_row}", rule)
    
    auto_fit_columns(ws, max_width=20)


def write_sell_through_sheet(writer, sell_through_df):
    from openpyxl.formatting.rule import ColorScaleRule
    sell_through_df.to_excel(writer, sheet_name="消化率・シーズン分析", index=False)
    ws = writer.sheets["消化率・シーズン分析"]
    header = [ws.cell(row=1,column=c).value for c in range(1,ws.max_column+1)]
    si = next((i+1 for i,h in enumerate(header) if h=="消化率"), None)
    if si:
        sl = ws.cell(row=1,column=si).column_letter
        for row in ws.iter_rows(min_row=2, min_col=si, max_col=si):
            for cell in row: cell.number_format = "0%"
        ws.conditional_formatting.add(f"{sl}2:{sl}5000", ColorScaleRule(
            start_type="num",start_value=0,start_color="FFC7CE",
            mid_type="num",mid_value=0.5,mid_color="FFEB9C",
            end_type="num",end_value=1,end_color="C6EFCE"))
    auto_fit_columns(ws)


def write_kanken_sheet(writer, df_merged, cols):
    cfg = get_config()
    kw = "|".join(cfg["analysis"]["kanken_keywords"])
    col_cat_m=cols["m_cat_m"]; col_store=cols["s_store"]
    col_color=cols["m_color"]; col_size=cols["s_size"]; col_reg=cols["s_reg"]
    col_key = cols["s_key"]

    df_k = df_merged[df_merged[col_cat_m].astype(str).str.contains(kw, case=False, na=False)]
    if df_k.empty: return

    df_23510 = df_k[df_k[col_key].astype(str).str.startswith("23510")]
    df_others = df_k[~df_k[col_key].astype(str).str.startswith("23510")]

    _write_kanken_section(writer, df_23510, "Kanken(23510)詳細", col_color, col_store, col_reg)
    _write_kanken_section(writer, df_others, "Kanken(その他)詳細", col_color, col_store, col_reg)


def _write_kanken_section(writer, df_target, sheet_name, col_color, col_store, col_reg):
    if df_target.empty: return

    # カラー全体の売上ランキング
    cs = df_target.groupby(col_color)[col_reg].count().sort_values(ascending=False).reset_index()
    cs.columns = ["カラー","販売数"]

    # 店舗ごとのTOP10
    store_color_counts = df_target.groupby([col_store, col_color])[col_reg].count().reset_index()
    store_color_counts.columns = ["店舗", "カラー", "販売数"]
    
    df_sorted = store_color_counts.sort_values(["店舗", "販売数"], ascending=[True, False])
    df_sorted["pos_rank"] = df_sorted.groupby("店舗").cumcount() + 1
    top10 = df_sorted[df_sorted["pos_rank"] <= 10].copy()
    top10["display"] = top10["カラー"] + " (" + top10["販売数"].astype(str) + ")"
    
    matrix = top10.pivot(index="pos_rank", columns="店舗", values="display").fillna("-")
    matrix.index.name = "ランキング"
    matrix = matrix.reset_index()

    cs.to_excel(writer, sheet_name=sheet_name, index=False)
    ws_k = writer.sheets[sheet_name]

    from openpyxl.formatting.rule import DataBarRule
    rule = DataBarRule(start_type="min", end_type="max", color="638EC6", showValue=True)
    ws_k.conditional_formatting.add(f"B2:B{len(cs)+1}", rule)

    sr = len(cs) + 3
    ws_k.cell(row=sr, column=1, value="【店舗別 売れ筋カラーTOP10】")
    matrix.to_excel(writer, sheet_name=sheet_name, startrow=sr, startcol=0, index=False)
    
    from utils import auto_fit_columns
    auto_fit_columns(ws_k)


def write_inbound_sheet(writer, df_merged, cols):
    cfg = get_config()
    kw = "|".join(cfg["column_mapping"]["inbound_keywords"])
    col_inbound=cols["s_inbound"]; col_cat_m=cols["m_cat_m"]
    col_price=cols["s_price"]; col_reg=cols["s_reg"]
    if not col_inbound: return

    df_merged["_ib"] = df_merged[col_inbound].astype(str).str.contains(kw, case=False, na=False)
    ib = df_merged[df_merged["_ib"]].groupby(col_cat_m).agg(
        販売数=(col_reg,"count"),
        売上金額=(col_price,"sum") if col_price else (col_reg, lambda x:0),
    ).sort_values("販売数", ascending=False)
    ib["構成比"] = ib["販売数"] / ib["販売数"].sum()
    ib.to_excel(writer, sheet_name="インバウンド分析")
    
    ws = writer.sheets["インバウンド分析"]
    header = [ws.cell(row=1,column=c).value for c in range(1,ws.max_column+1)]
    if "構成比" in header:
        ci = header.index("構成比") + 1
        cl = ws.cell(row=1,column=ci).column_letter
        for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
            for cell in row: cell.number_format = "0%"
            
        from openpyxl.formatting.rule import DataBarRule
        ws.conditional_formatting.add(f"{cl}2:{cl}{ws.max_row}",
            DataBarRule(start_type="min", end_type="max", color="638EC6", showValue=True))
            
    auto_fit_columns(ws)


def write_yoy_sheet(writer, yoy_df):
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill, Font
    if yoy_df is None or yoy_df.empty: return

    # 全体合計を計算して追加
    total_yoy = yoy_df.groupby("年月").agg({
        "今年販売数": "sum", "今年売上金額": "sum",
        "前年販売数": "sum", "前年売上金額": "sum"
    }).reset_index()
    total_yoy["中分類"] = "【全体合計】"
    total_yoy["販売数前年比"] = total_yoy.apply(lambda r: r["今年販売数"] / r["前年販売数"] if r["前年販売数"] > 0 else None, axis=1)
    total_yoy["売上金額前年比"] = total_yoy.apply(lambda r: r["今年売上金額"] / r["前年売上金額"] if r["前年売上金額"] > 0 else None, axis=1)
    
    # 元データに結合
    yoy_df = pd.concat([total_yoy, yoy_df], ignore_index=True)
    
    # 縦並びのデータをマトリクス（横並び）にピボット
    pivot_qty = yoy_df.pivot_table(index="中分類", columns="年月", values="販売数前年比", aggfunc="mean").fillna(0)
    pivot_qty.columns.name = None
    pivot_qty.index.name = "中分類（販売数前年比）"
    pivot_qty = pivot_qty.reset_index()

    # 「【全体合計】」を一番上にするためにカスタムソート
    pivot_qty["_sort"] = pivot_qty["中分類（販売数前年比）"].apply(lambda x: 0 if x == "【全体合計】" else 1)
    pivot_qty = pivot_qty.sort_values(["_sort", "中分類（販売数前年比）"]).drop(columns=["_sort"]).reset_index(drop=True)

    pivot_amt = yoy_df.pivot_table(index="中分類", columns="年月", values="売上金額前年比", aggfunc="mean").fillna(0)
    pivot_amt.columns.name = None
    pivot_amt.index.name = "中分類（売上金額前年比）"
    pivot_amt = pivot_amt.reset_index()
    pivot_amt["_sort"] = pivot_amt["中分類（売上金額前年比）"].apply(lambda x: 0 if x == "【全体合計】" else 1)
    pivot_amt = pivot_amt.sort_values(["_sort", "中分類（売上金額前年比）"]).drop(columns=["_sort"]).reset_index(drop=True)

    # シートへ書き出し
    pivot_qty.to_excel(writer, sheet_name="前年同月比", startrow=0, index=False)
    
    start_row_amt = len(pivot_qty) + 3
    pivot_amt.to_excel(writer, sheet_name="前年同月比", startrow=start_row_amt, index=False)
    
    ws = writer.sheets["前年同月比"]
    
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100")
    
    from openpyxl.utils import get_column_letter
    max_col = ws.max_column
    
    # 条件付き書式と%表示（表1：販売数）
    for col in range(2, max_col + 1):
        col_letter = get_column_letter(col)
        for row in range(2, len(pivot_qty) + 2):
            ws[f"{col_letter}{row}"].number_format = "0%"
        rng1 = f"{col_letter}2:{col_letter}{len(pivot_qty) + 1}"
        ws.conditional_formatting.add(rng1, CellIsRule(operator="lessThan", formula=["1.0"], fill=red_fill, font=red_font))
        ws.conditional_formatting.add(rng1, CellIsRule(operator="greaterThanOrEqual", formula=["1.0"], fill=green_fill, font=green_font))

    # 条件付き書式と%表示（表2：売上金額）
    start_r = start_row_amt + 2
    end_r = start_row_amt + len(pivot_amt) + 1
    for col in range(2, max_col + 1):
        col_letter = get_column_letter(col)
        for row in range(start_r, end_r + 1):
            ws[f"{col_letter}{row}"].number_format = "0%"
        rng2 = f"{col_letter}{start_r}:{col_letter}{end_r}"
        ws.conditional_formatting.add(rng2, CellIsRule(operator="lessThan", formula=["1.0"], fill=red_fill, font=red_font))
        ws.conditional_formatting.add(rng2, CellIsRule(operator="greaterThanOrEqual", formula=["1.0"], fill=green_fill, font=green_font))

    auto_fit_columns(ws)

    # 折れ線グラフ用：「全体合計」のデータを縦方向に転置（0%=データなし月は除外）
    from openpyxl.chart import LineChart, Reference as ChartRef
    month_cols = [c for c in pivot_qty.columns if c != "中分類（販売数前年比）"]
    total_row = pivot_qty[pivot_qty["中分類（販売数前年比）"] == "【全体合計】"]
    
    if not total_row.empty and len(month_cols) > 0:
        # データがある月だけ抽出（0%の未来月を除外）
        valid_months = []
        for m in month_cols:
            val = total_row[m].values[0]
            if val and val > 0:
                valid_months.append((m, val))
        
        if len(valid_months) > 0:
            chart_data_start = end_r + 3
            # ヘッダー行
            ws.cell(row=chart_data_start, column=1, value="年月")
            ws.cell(row=chart_data_start, column=2, value="全体 販売数前年比")
            ws.cell(row=chart_data_start, column=3, value="100%ライン")
            
            for i, (m, val) in enumerate(valid_months):
                r = chart_data_start + 1 + i
                ws.cell(row=r, column=1, value=m)
                ws.cell(row=r, column=2, value=val)
                ws.cell(row=r, column=2).number_format = "0%"
                ws.cell(row=r, column=3, value=1.0)
                ws.cell(row=r, column=3).number_format = "0%"
            
            chart_data_end = chart_data_start + len(valid_months)
            
            line = LineChart()
            line.title = "ブランド全体 前年同月比推移（販売数）"
            line.style = 10
            line.y_axis.title = "前年比"
            line.x_axis.title = "年月"
            line.y_axis.numFmt = "0%"
            
            # B列：前年比データ
            data = ChartRef(ws, min_col=2, min_row=chart_data_start, max_row=chart_data_end)
            line.add_data(data, titles_from_data=True)
            # C列：100%基準線
            baseline = ChartRef(ws, min_col=3, min_row=chart_data_start, max_row=chart_data_end)
            line.add_data(baseline, titles_from_data=True)
            # A列：年月ラベル
            cats = ChartRef(ws, min_col=1, min_row=chart_data_start + 1, max_row=chart_data_end)
            line.set_categories(cats)
            
            # 100%ラインを灰色の点線にする
            s2 = line.series[1]
            s2.graphicalProperties.line.solidFill = "AAAAAA"
            s2.graphicalProperties.line.dashStyle = "dash"
            s2.graphicalProperties.line.width = 12700  # 1pt
            
            line.height = 15; line.width = 25
            ws.add_chart(line, f"D{chart_data_start}")


def write_stock_health_sheet(writer, trend, df_merged, cols):
    """
    在庫ヘルス推移シートを追加し、過去の在庫数量・金額・WOS推移テーブルとグラフを描画する
    """
    if trend is None or trend.empty:
        # 初回実行時などの空状態へのフォールバック
        ws = writer.book.create_sheet(title="在庫ヘルス推移")
        ws["A1"] = "在庫ヘルス推移レポート (過去8週間)"
        ws["A1"].font = ws["A1"].font.copy(bold=True, size=14)
        ws["A3"] = "【お知らせ】過去の在庫履歴がまだデータベースに十分に蓄積されていません。"
        ws["A4"] = "このツールを実行するたびに、その日の在庫状況が自動的にDBにスナップショット保存されます。"
        ws["A5"] = "次回（翌日や来週など）以降の実行から、このシートに時系列グラフが自動描画されます。"
        return

    # 中分類ごとの最新平均週販数を取得 (df_mergedから)
    col_cat_m = cols["m_cat_m"]
    col_key = cols["s_key"]
    # df_merged内のweekly_avgは品番単位なので、中分類ごとに品番の週販を合計する
    cat_sales_map = df_merged.drop_duplicates(subset=[col_key]).groupby(col_cat_m)["weekly_avg"].sum()
    
    # trend(中分類, 実行日) に対して WOS を計算して列を追加
    def calc_row_wos(row):
        cat = row["中分類"]
        qty = row["在庫数量合計"]
        weekly_sales = cat_sales_map.get(cat, 0.01)
        if weekly_sales <= 0.01:
            return 99.9  # 週販がほぼない場合は安全上限値
        return round(qty / weekly_sales, 1)

    trend["WOS"] = trend.apply(calc_row_wos, axis=1)

    # ピボットテーブルの作成（横：実行日、縦：中分類）
    pivot_qty = trend.pivot_table(index="中分類", columns="実行日", values="在庫数量合計", aggfunc="sum").fillna(0)
    pivot_qty.columns.name = None
    pivot_qty = pivot_qty.reset_index()

    pivot_amt = trend.pivot_table(index="中分類", columns="実行日", values="在庫金額合計", aggfunc="sum").fillna(0)
    pivot_amt.columns.name = None
    pivot_amt = pivot_amt.reset_index()

    pivot_wos = trend.pivot_table(index="中分類", columns="実行日", values="WOS", aggfunc="mean").fillna(0)
    pivot_wos.columns.name = None
    pivot_wos = pivot_wos.reset_index()

    # Excelへの書き出し
    pivot_qty.to_excel(writer, sheet_name="在庫ヘルス推移", startrow=4, index=False)
    
    start_row_amt = len(pivot_qty) + 8
    pivot_amt.to_excel(writer, sheet_name="在庫ヘルス推移", startrow=start_row_amt, index=False)

    start_row_wos = start_row_amt + len(pivot_amt) + 4
    pivot_wos.to_excel(writer, sheet_name="在庫ヘルス推移", startrow=start_row_wos, index=False)

    ws = writer.sheets["在庫ヘルス推移"]
    ws["A1"] = "在庫ヘルス推移レポート (過去8週間)"
    ws["A1"].font = ws["A1"].font.copy(bold=True, size=14)
    ws["A2"] = "※このシートはSQLite DBに蓄積された在庫スナップショット履歴から自動的に生成されます。"
    
    ws.cell(row=4, column=1, value="中分類（在庫数量）")
    ws.cell(row=4, column=1).font = ws.cell(row=4, column=1).font.copy(bold=True)
    
    # 金額のテーブル書き出し
    ws.cell(row=start_row_amt - 1, column=1, value="中分類（在庫金額：原価）")
    ws.cell(row=start_row_amt - 1, column=1).font = ws.cell(row=start_row_amt - 1, column=1).font.copy(bold=True)

    # WOSのテーブル書き出し
    ws.cell(row=start_row_wos - 1, column=1, value="中分類（適正在庫WOS：最新週販換算）")
    ws.cell(row=start_row_wos - 1, column=1).font = ws.cell(row=start_row_wos - 1, column=1).font.copy(bold=True)

    # フォーマットを適用
    max_col = ws.max_column
    for col in range(2, max_col + 1):
        for row in range(start_row_amt + 1, start_row_amt + len(pivot_amt) + 1):
            ws.cell(row=row, column=col).number_format = "¥#,##0"
        for row in range(start_row_wos + 1, start_row_wos + len(pivot_wos) + 1):
            ws.cell(row=row, column=col).number_format = "0.0"

    # 共通のカテゴリー軸（x軸）: 4行目のB列以降（実行日の日付ヘッダー）
    from openpyxl.chart import Reference as ChartRef
    cats = ChartRef(ws, min_col=2, min_row=4, max_col=max_col, max_row=4)

    # --- 1. 在庫数量推移 折れ線グラフ ---
    from openpyxl.chart import LineChart
    line_qty = LineChart()
    line_qty.title = "週次 在庫数量推移（中分類別）"
    line_qty.style = 10
    line_qty.y_axis.title = "在庫数量"
    line_qty.x_axis.title = "実行日"
    
    data_qty = ChartRef(ws, min_col=1, min_row=5, max_col=max_col, max_row=4 + len(pivot_qty))
    line_qty.add_data(data_qty, titles_from_data=True, from_rows=True)
    line_qty.set_categories(cats)
    line_qty.height = 10; line_qty.width = 16
    ws.add_chart(line_qty, f"B{len(pivot_qty) + 6}")

    # --- 2. 在庫金額推移 積み上げ棒グラフ ---
    from openpyxl.chart import BarChart
    bar = BarChart()
    bar.type = "col"
    bar.style = 11
    bar.grouping = "stacked"
    bar.overlap = 100
    bar.title = "週次 在庫金額推移（中分類別・原価）"
    bar.y_axis.title = "在庫金額"
    bar.x_axis.title = "実行日"
    bar.y_axis.numFmt = "¥#,##0"
    
    data_amt = ChartRef(ws, min_col=1, min_row=start_row_amt + 1, max_col=max_col, max_row=start_row_amt + len(pivot_amt))
    bar.add_data(data_amt, titles_from_data=True, from_rows=True)
    bar.set_categories(cats)
    bar.height = 10; bar.width = 16
    ws.add_chart(bar, f"B{start_row_amt + len(pivot_amt) + 2}")

    # --- 3. 在庫週数(WOS)推移 折れ線グラフ ---
    line_wos = LineChart()
    line_wos.title = "週次 在庫週数(WOS)推移（中分類別・最新週販換算）"
    line_wos.style = 13
    line_wos.y_axis.title = "在庫週数 (WOS)"
    line_wos.x_axis.title = "実行日"
    
    data_wos = ChartRef(ws, min_col=1, min_row=start_row_wos + 1, max_col=max_col, max_row=start_row_wos + len(pivot_wos))
    line_wos.add_data(data_wos, titles_from_data=True, from_rows=True)
    line_wos.set_categories(cats)
    line_wos.height = 10; line_wos.width = 16
    ws.add_chart(line_wos, f"B{start_row_wos + len(pivot_wos) + 2}")

    auto_fit_columns(ws)


